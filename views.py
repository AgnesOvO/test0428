from flask import Blueprint, render_template, request, flash, jsonify, redirect
from sqlalchemy.sql.functions import user
from flask_login import login_required, current_user
from .models import Note
from . import db
import json
import os
import random
import math
import googletrans
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.styles import Font
import pandas as pd
from PIL import Image
from googletrans import Translator
import numpy as np
from tkinter import messagebox

views = Blueprint('views', __name__)

# 上傳檔案
@views.route('/', methods=["GET", "POST"])
@login_required
def home():
        
    if request.method == "POST":

        if request.files:

            xlsx = request.files["xlsx"]

            if xlsx.filename == "":
                flash('Fail to upload!', category='fail')
                return redirect(request.url)

            upload_path = os.path.join(
                os.path.join(os.environ['USERPROFILE']), 'Desktop')

            xlsx.save(upload_path)

            str_upload_path = str(upload_path)

            os.rename(str_upload_path + xlsx.filename,
                      str_upload_path + "/ori.xlsx")

            flash('File Uploaded!', category='succes')
            return render_template("home.html", user=current_user)

    return render_template("home.html", user=current_user)

#輸入欄位名稱_藏入圖片
@views.route('/entercolname-pic', methods=["POST"])
@login_required
def entercolname_pic():
    return render_template("entercolname_pic.html", user=current_user)

# 藏入圖片
@views.route("/picture", methods=["GET"])
@login_required
def addpicture():

    upload_path = os.path.join(
                os.path.join(os.environ['USERPROFILE']), 'Desktop')

    # 使用openpyxl建立新活頁簿wb_new
    wb_new = Workbook()
    wb_new.save(upload_path + '/new_excel_pic.xlsx')

    # 使用openpyxl讀取原始檔案
    wb = load_workbook(upload_path + '/ori.xlsx')
    ws = wb.worksheets[0]

    # 使用openpyxl讀取new_excel
    wb_new = load_workbook(upload_path + '/new_excel_pic.xlsx')
    ws_new = wb_new.active

    # 使用openpyxl計算原始excel欄列總數
    max_row_wb = ws.max_row
    max_col_wb = ws.max_column

    # 建立listab
    listab = []
    for k in range(1, max_row_wb + 1):
        listab.append([])

    # 複製原始excel的cell
    for r in range(1, max_row_wb + 1):
        for c in range(1, max_col_wb + 1):
            e = ws.cell(row=r, column=c)
            listab[r-1].append(e.value)

    # 貼上new_excel
    for r in range(1, max_row_wb + 1):
        for c in range(1, max_col_wb + 1):
            m = ws_new.cell(row=r, column=c)
            m.value = listab[r-1][c-1]

    wb_new.save(upload_path + '/new_excel_pic.xlsx')

    #使用pd讀取原始excel
    df = pd.read_excel(upload_path + '/ori.xlsx')
    name = request.args.get('piccolname')
    #df = pd.read_excel("32.xlsx")
    score=df[name]

    #a = score長度
    a = len(score)
    #print('指定欄位長度:', a)
    List1 = []
    if a % 2 ==0 :
    #存a到List陣列
        for i in range(0, a):
            result = 0
            result = score[i]
            List1.insert(i, result)
        #print("原始資料:",List1)
        list1=len(List1)
    #多補一個0到最後
    elif a % 2 != 0 :
        for i in range(0, a+1):
            if i < a :
                result = 0
                result = score[i]
                List1.insert(i, result)
            elif  i == a :
                result = 0
                List1.insert(i,result)
        list1=len(List1)
        #print("原始資料:",List1)
        #print("補0後的長度:",list1)
    hide1_len = int(list1/2)

    ####準備第一次擴張
    Listd = []

    Listm = []

    Listd2 = []
    for i in range (0,list1,2) :

            k = i//2
            act = List1[i]+List1[i+1]
            d = abs(List1[i]-List1[i+1])
            m = act//2
            Listd.insert(k,d)
            Listm.insert(k,m)
            dl = len(Listd)
    #print("d:",Listd)
    #print("d的長度:",dl)
    #print("m:",Listm)

    #藏入1的數量
    len2 = 2*hide1_len
    #print("藏入1的數量",len2)

    for i in range (0,dl) :

                d2 = 2*Listd[i]+1
                Listd2.insert(i,d2)
    ##擴張
    List_single = []

    for i in range (0,dl) :
            if List1[i*2] >= List1[i*2+1] :
                result1 = Listm[i]+((Listd2[i]+1)//2)
                result2 = Listm[i]-(Listd2[i]//2)
                m = (result1+result2)//2
                d2 = (abs(result1-result2)*2)+1
                resulta = m+((d2+1)//2)
                resultb = m-(d2//2)
                
            else :
                result1 =Listm[i]-(Listd2[i]//2)
                result2 =Listm[i]+((Listd2[i]+1)//2)
                m = (result1+result2)//2
                d2 = (abs(result1-result2)*2)+1
                resulta = m-(d2//2)
                resultb = m+((d2+1)//2)

            List_single.insert(i*2,resulta)
            List_single.insert(i*2+1,resultb)
    len_single = len(List_single)
    #print("每組藏入2bit後資料:",List_single)

    #溢位處理
    List_over_1 = []
    List_lebel_1 = []
    for i in range (0,len_single) :

        k = List_single[i]
        
        if k < 0 :
            result_1_underflow =abs(k)
            List_over_1.insert(i,result_1_underflow)
            List_lebel_1.insert(i,1)


        else :
            List_over_1.insert(i,k)
            List_lebel_1.insert(i,0)
    listover_1 = len(List_over_1)
    listlebel_1 = len(List_lebel_1)

    print("藏入全白圖片(變造後)之資料:",List_over_1)
    #print("哪些處理過(1為有處理):",List_lebel_1)

    # 讀取new_excel
    wb_new = load_workbook(upload_path + '/new_excel_pic.xlsx')
    ws_new = wb_new.active

    # 建立所有欄位名稱的List，取得指定欄位的位置
    List_colName = []
    List_colName = [item.value for item in list(ws.rows)[0]]
    collen = len(List_colName)
    #print('檔案欄位總數:', collen)

    str1 = name

    for x in range(1, collen):
        if List_colName[x] == str1:
            # 欄位名稱與指定欄位相等
            c = x + 1
            # c=指定欄位的位置

    #print(str(name)+'位於:第', c,'欄')

    # 使用openpyxl寫入變造後資料
    i = 0
    j = 0
    for r in range(2, listover_1+2):
        k_1 = List_lebel_1[j]
        if k_1 ==1:
            font1 = Font(size=11.5)
            ws_new.cell(row=r, column=c, value=List_over_1[i]).font=font1

        else :
            font2 = Font(size=12)
            ws_new.cell(row=r, column=c, value=List_over_1[i]).font=font2
        i += 1
        j += 1

    # 儲存檔案
    wb_new.save(upload_path + '/new_excel_pic.xlsx')

    if os.path.exists(upload_path + "/ori.xlsx"):
        os.remove(upload_path + "/ori.xlsx")

    return redirect("/")

#輸入欄位名稱_取出圖片
@views.route('/entercolname-takeoutpic', methods=["POST"])
@login_required
def entercolname_takeoutpic():
    return render_template("entercolname_takeoutpic.html", user = current_user)

# 取出圖片
@views.route("/takeoutpicture", methods=["GET"])
@login_required
def takeoutpicture():

    upload_path = os.path.join(
                os.path.join(os.environ['USERPROFILE']), 'Desktop') 
    # 使用openpyxl建立新活頁簿wb_new
    #wb_new = Workbook()
    #wb_new.save('c:/Users/NKUST/Desktop/presentation/ori_excel_pic.xlsx')

    # 使用openpyxl讀取原始檔案
    wb = load_workbook(upload_path + '/ori.xlsx')
    ws = wb.worksheets[0]


    # 使用openpyxl讀取new_excel
    #wb_new = load_workbook('c:/Users/NKUST/Desktop/presentation/ori_excel_pic.xlsx')
    #ws_new = wb_new.active

    # 使用openpyxl計算原始excel欄列總數
    max_row_wb = ws.max_row
    max_col_wb = ws.max_column
    #print(max_row_wb)
    #print(max_col_wb)


    # 建立listab
    listab = []
    for k in range(1, max_row_wb + 1):
        listab.append([])

    # 複製原始excel的cell
    for r in range(1, max_row_wb + 1):
        for c in range(1, max_col_wb + 1):
            e = ws.cell(row=r, column=c)
            listab[r-1].append(e.value)

    # 貼上new_excel
    #for r in range(1, max_row_wb + 1):
        #for c in range(1, max_col_wb + 1):
            #m = ws_new.cell(row=r, column=c)
            #m.value = listab[r-1][c-1]
            #print(m)


    #wb_new.save('c:/Users/NKUST/Desktop/presentation/ori_excel_pic.xlsx')

    #使用pd讀取原始excel
    df = pd.read_excel(upload_path + '/ori.xlsx')
    name = request.args.get('takeoutpic-colname')
    #df = pd.read_excel("new_22.xlsx")
    score=df[name]

    # 建立所有欄位名稱的List，取得指定欄位的位置
    List_colName = []
    List_colName = [item.value for item in list(ws.rows)[0]]
    collen = len(List_colName)
    #print('欄位總數:', collen)
    str1 = name

    for x in range(1, collen):
        if List_colName[x] == str1:
            # 欄位名稱與指定欄位相等
            c = x + 1
            # c=指定欄位的位置

    #a = score長度
    a = len(score)
    #print('指定欄位長度:', a)

    if a %2 !=0 :  #若資料長度為奇數
        #存a到List2陣列
        List2 = []
        for i in range(0, a):
            result = score[i]
            List2.insert(i, result)
        List2.insert(a,0)
        #print("之前變造完成的資料:",List2)
    else :
        List2 = []
        for i in range(0, a):
            result = score[i]
            List2.insert(i, result)
        #print("之前變造完成的資料:",List2)

    #存處理前資料到List_foward陣列
    List_foward = []
    Listp1 = []


    fontt = []
    r=2
    try:
        for j in range(0,max_row_wb):
            cell = ws.cell(row=r, column=c)
            font = cell.font.size
            if font == 11.5 :
                Listp1.insert(j,1)
                fontt.insert(j,font)
                List_foward.insert(j,-List2[j])
            else:
                Listp1.insert(j,0)
                fontt.insert(j,font)
                List_foward.insert(j,List2[j])  
            r+=1
    except IndexError:
        pass
    #print("FONT",fontt)
    #print("處理前的資料:",List_foward)
    #print("over:",Listp1)

    Listd2d = []

    Listm2m = []

    #取出d'跟 m'(還原資料需要的一部分步驟)
    for i in range (0,a,2) :

        k = i//2

        act = List_foward[i]+List_foward[i+1]

        d2d = abs( List_foward[i]- List_foward[i+1])
        m2m = act//2
        Listd2d.insert(k,d2d)
        Listm2m.insert(k,m2m)
    #print("D:",Listd2d)
    #print("m",Listm2m)
    #print("d2d:",Listd2d)
    d2dl = len(Listd2d)
    #print("d2dl長度:",d2dl)
    #print("m2m:",Listm2m)


    ListS= []

    for i in range (0,d2dl) :
        #b = (i*2)+1
        s2 = Listd2d[i] % 2
        ListS.insert(i,s2)
        #ListBool.insert(b,s)
    #print("B1LIST",ListBool)
    #求出擴張一次原始d
    Listd = []
    for i in range (0,d2dl) :
        d = (Listd2d[i]-ListS[i])//2
        Listd.insert(i,d)
    #print("原始d:",Listd)

    #找出圖片訊息
    ListBool = []
    Listori = []
    for i in range (0,d2dl) :
        if List_foward[i*2] >= List_foward[i*2+1] :
            a = Listm2m[i] + ((Listd[i]+1)//2)
            b = Listm2m[i] - (Listd[i]//2)
            m = (a+b)//2
            d2 = abs(a-b)
            s = d2%2
            d = (d2-s)//2
            result = m + ((d+1)//2)
            result2 = m - (d//2)   
            ListBool.insert(i*2,s)
            ListBool.insert(i*2+1,ListS[i])
            Listori.insert(i*2,result)
            Listori.insert(i*2+1,result2)
        else :
            a = Listm2m[i] - (Listd[i]//2)
            b = Listm2m[i] + ((Listd[i]+1)//2)
            m = (a+b)//2
            d2 = abs(a-b)
            s = d2%2
            d = (d2-s)//2
            result = m - (d//2)
            result2 = m + ((d+1)//2)  
            ListBool.insert(i*2,s)
            ListBool.insert(i*2+1,ListS[i])
            Listori.insert(i*2,result)
            Listori.insert(i*2+1,result2)
    listbool_len = len(ListBool)
    listori_len = len(Listori)
    #print("原始資料:",Listori)
    #print("BOOL2",ListBool)
    #print("BOOLEAN長度",listbool_len)
    #print("LISTBOO",ListBool)

    #圖片部分
    p = round(math.sqrt((listbool_len)))
    flash("圖片大小:"+str(p)+"*"+str(p))
    p2 = p**2
    #print(p2)

    if p2 < listbool_len :
        less = listbool_len - p2
        #print("提取訊息之末",less,"碼，未被採用成像素")
    if p2 > listbool_len :
        support = p2 - listbool_len
        for i in range(0,support) :
            ListBool.insert(listbool_len,1)
            listbool_len += 1
        listbool_len=len(ListBool)
        #print("多補了",support,"個1 當純白像素")
    #print("補1後 BOOLEANLIST:",ListBool)


    List_wrong = []
    h2 = [[0]*p for i in range(p)]
    i = 0 
    w=1
    for j in range (0,p) :
        a = 0
    
        while a < p :
            t = ListBool[i]+ListBool[i+1]
            if t == 2 :
                h2[j][a] = True
                h2[j][a+1] = True
            else :
                h2[j][a] = False
                h2[j][a+1] = False
                List_wrong.append(w)
                List_wrong.append(w+1)
            w+=2
            a+=2
            i+=2
    listwrong=len(List_wrong)
    #print(listwrong)
    group_wrong = listwrong//2
    h2len = len(h2)

    a1 = len(score)

    #錯誤率
    wrong_persent = round((group_wrong/5*100),2)
    accept = 100.00-wrong_persent
    accept = format(accept,'.2f')

    #print("圖片內pixel值:\n",h2)

    
    if listwrong == 0 :
        print("無遭竄改，資料正確性高")
    else :
        flash("第"+str(List_wrong)+"筆資料可能遭更動")
        flash("約有"+str(group_wrong)+"組錯誤")
        #print("真正錯誤組數:",5)
        #print("實際偵測出錯誤組數",group_wrong)
    #    print("錯誤偵測率 約:",wrong_persent,"%")
    #    print("錯誤接受率 約:",accept,"%")


    #print(h2len)



    mats = np.array(h2)
    mg2 = Image.fromarray(mats)
    mg2.save(upload_path + "/BW2.png")

    #檢查List_new長度
    listori_len = len(Listori)
    #print("原始資料長度:", Listori_len)

    if os.path.exists(upload_path + "/ori.xlsx"):
        os.remove(upload_path + "/ori.xlsx")


    return redirect("/")